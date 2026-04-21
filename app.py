import io
import re
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from extractor import extract_commission_data


def _md(html: str) -> None:
    """Render HTML via st.markdown — collapses whitespace so markdown
    never misinterprets indented lines as code blocks."""
    st.markdown(re.sub(r"\s+", " ", html).strip(), unsafe_allow_html=True)

COLUMNS = [
    "Solicitud de Comisión No.",
    "Nombre",
    "Tipo de Documento",
    "Número de Documento",
    "Cargo",
    "Fecha Inicial Comisión",
    "Fecha Final Comisión",
    "Dpto. / Municipio Origen",
    "Dpto. / Municipio Destino",
    "Valor Total a Pagar",
    "Objeto de la Comisión",
]

EARM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;700&display=swap');

/* ── Global ── */
html, body, .stApp, [data-testid="stAppViewContainer"] {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    background-color: #FAFAFA !important;
    color: #424242 !important;
}

/* Orbes ambientales */
[data-testid="stAppViewContainer"]::before {
    content: '';
    position: fixed;
    top: -18%; left: -8%;
    width: 384px; height: 384px;
    background: rgba(255,238,88,0.18);
    border-radius: 50%;
    filter: blur(120px);
    pointer-events: none;
    z-index: 0;
}
[data-testid="stAppViewContainer"]::after {
    content: '';
    position: fixed;
    bottom: -18%; right: -8%;
    width: 384px; height: 384px;
    background: rgba(16,185,129,0.13);
    border-radius: 50%;
    filter: blur(120px);
    pointer-events: none;
    z-index: 0;
}

/* ── Toolbar Streamlit (barra superior) ── */
[data-testid="stHeader"] {
    background: rgba(44,44,44,0.97) !important;
    backdrop-filter: blur(20px) !important;
    border-bottom: 2.5px solid #FFEE58 !important;
}
[data-testid="stToolbar"] { display: none !important; }

/* ── Bloque de contenido ── */
.block-container {
    padding-top: 1.5rem !important;
    padding-bottom: 3rem !important;
    max-width: 1200px !important;
}

/* ── Tipografía ── */
h1, h2, h3, h4 {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 800 !important;
    color: #2C2C2C !important;
}
p, span, div, label, td, th {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.80) !important;
    backdrop-filter: blur(20px) !important;
    border-radius: 16px !important;
    border: 2px dashed #FFEE58 !important;
    padding: 0.5rem !important;
    transition: all 0.2s !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: transparent !important;
    border: none !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: #FDD835 !important;
    background: rgba(255,238,88,0.05) !important;
    box-shadow: 0 0 0 4px rgba(255,238,88,0.12) !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] {
    color: #424242 !important;
}

/* ── Botones ── */
.stButton > button,
[data-testid="stDownloadButton"] > button {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 700 !important;
    background: #FFEE58 !important;
    color: #2C2C2C !important;
    border: none !important;
    border-radius: 12px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    font-size: 0.78rem !important;
    padding: 0.55rem 1.25rem !important;
    transition: all 0.2s !important;
    box-shadow: 0 2px 8px rgba(255,238,88,0.25) !important;
}
.stButton > button:hover,
[data-testid="stDownloadButton"] > button:hover {
    background: #FDD835 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 16px rgba(255,238,88,0.40) !important;
}

/* ── Alertas ── */
[data-testid="stAlert"] {
    border-radius: 12px !important;
    backdrop-filter: blur(12px) !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 500 !important;
}

/* ── Barra de progreso ── */
[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #FFEE58, #FDD835) !important;
    border-radius: 8px !important;
}
[data-testid="stProgressBar"] > div {
    background: #F0F0F0 !important;
    border-radius: 8px !important;
}

/* ── DataFrames ── */
[data-testid="stDataFrame"] {
    background: rgba(255,255,255,0.80) !important;
    backdrop-filter: blur(20px) !important;
    border-radius: 16px !important;
    border: 1px solid #F0F0F0 !important;
    overflow: hidden !important;
    box-shadow: 0 2px 12px rgba(0,0,0,0.04) !important;
}

/* ── Expander (advertencias) ── */
[data-testid="stExpander"] {
    background: rgba(255,255,255,0.80) !important;
    backdrop-filter: blur(20px) !important;
    border-radius: 16px !important;
    border: 1px solid #F0F0F0 !important;
}

/* ── Info box ── */
[data-testid="stInfo"] {
    background: rgba(239,246,255,0.90) !important;
    border-left: 4px solid #3b82f6 !important;
    border-radius: 12px !important;
}

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: #F0F0F0; }
::-webkit-scrollbar-thumb { background: #ccc; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #FFEE58; }
</style>
"""

HEADER_HTML = """
<div style="
    background: linear-gradient(135deg, #2C2C2C 0%, #1a1a1a 100%);
    border-bottom: 3px solid #FFEE58;
    border-radius: 0 0 20px 20px;
    padding: 1.25rem 1.75rem;
    margin-bottom: 1.75rem;
    display: flex;
    align-items: center;
    gap: 1rem;
    box-shadow: 0 8px 32px rgba(0,0,0,0.18);
    position: relative;
    overflow: hidden;
">
    <div style="position:absolute;top:-40px;right:-40px;width:200px;height:200px;
        background:rgba(255,238,88,0.08);border-radius:50%;filter:blur(40px);"></div>

    <div style="
        width: 52px; height: 52px;
        background: #FFEE58;
        border-radius: 15px;
        display: flex; align-items: center; justify-content: center;
        font-size: 1.6rem;
        box-shadow: 0 4px 16px rgba(255,238,88,0.45);
        flex-shrink: 0;
    ">📄</div>

    <div style="flex: 1;">
        <div style="display: flex; align-items: center; gap: 10px; margin-bottom: 2px;">
            <span style="
                color: #FFFFFF;
                font-family: 'Plus Jakarta Sans', sans-serif;
                font-weight: 800;
                font-size: 1.2rem;
                letter-spacing: -0.02em;
            ">EXTRACTOR SIIF</span>
            <span style="
                font-family: 'JetBrains Mono', monospace;
                font-size: 9px;
                font-weight: 700;
                background: #F0F0F0;
                color: rgba(66,66,66,0.55);
                padding: 2px 7px;
                border-radius: 5px;
                letter-spacing: 0.1em;
            ">v1.0</span>
        </div>
        <span style="
            color: rgba(255,255,255,0.45);
            font-family: 'Plus Jakarta Sans', sans-serif;
            font-size: 0.72rem;
            letter-spacing: 0.01em;
        "></span>
    </div>

    <div style="display: flex; align-items: center; gap: 7px; flex-shrink: 0;">
        <div style="
            width: 8px; height: 8px;
            background: #10b981;
            border-radius: 50%;
            box-shadow: 0 0 0 3px rgba(16,185,129,0.25);
        "></div>
        <span style="
            color: #10b981;
            font-family: 'Plus Jakarta Sans', sans-serif;
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 0.15em;
            text-transform: uppercase;
        ">En línea</span>
    </div>
</div>
"""

FOOTER_HTML = """
<div style="
    margin-top: 3rem;
    border-top: 2px solid #F0F0F0;
    padding: 1rem 0 0.5rem;
    text-align: center;
">
    <span style="
        font-family: 'Plus Jakarta Sans', sans-serif;
        font-size: 0.72rem;
        font-weight: 700;
        color: rgba(66,66,66,0.55);
        letter-spacing: 0.12em;
        text-transform: uppercase;
    ">
        Herramienta desarrollada por el <strong style="color:#424242;">Despacho EARM</strong>
        &nbsp;&mdash;&nbsp;
        Derechos reservados &copy; <strong style="color:#424242;">CNSC</strong>
        &nbsp;&mdash;&nbsp; Edwin Arturo Ruiz Moreno &mdash; 2026
    </span>
</div>
"""

SECTION_STYLE = (
    "background: rgba(255,255,255,0.80);"
    "backdrop-filter: blur(20px);"
    "border: 1px solid #F0F0F0;"
    "border-radius: 16px;"
    "padding: 1.25rem 1.5rem;"
    "margin-bottom: 1.25rem;"
    "box-shadow: 0 2px 12px rgba(0,0,0,0.04);"
)


def section_header(icon, title, subtitle=""):
    color_map = {"📊": "#f59e0b", "✅": "#10b981", "⚠️": "#f59e0b", "📋": "#3b82f6"}
    color = color_map.get(icon, "#FFEE58")
    sub_html = f'<p style="font-size:0.72rem;color:rgba(66,66,66,0.55);margin:0;">{subtitle}</p>' if subtitle else ""
    return f"""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:1rem;">
        <div style="width:40px;height:40px;background:{color}22;border-radius:12px;
            display:flex;align-items:center;justify-content:center;font-size:1.1rem;">{icon}</div>
        <div>
            <h3 style="margin:0;font-size:1rem;font-weight:700;color:#2C2C2C;
                font-family:'Plus Jakarta Sans',sans-serif;">{title}</h3>
            {sub_html}
        </div>
    </div>
    """


def build_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comisiones"

    header_fill = PatternFill("solid", fgColor="2C2C2C")
    accent_fill = PatternFill("solid", fgColor="FFEE58")
    alt_fill    = PatternFill("solid", fgColor="F9F9F6")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            if df.columns[col_idx - 1] == "Valor Total a Pagar":
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = left

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    st.set_page_config(
        page_title="Extractor SIIF — Despacho EARM — CNSC",
        page_icon="📄",
        layout="wide",
    )

    _md(EARM_CSS)
    _md(HEADER_HTML)

    # ── Uploader ────────────────────────────────────────────────────────────
    _md(f'<div style="{SECTION_STYLE}">'
        + section_header("📂", "Cargar archivos PDF",
                         "Puede cargar múltiples PDFs simultáneamente — hasta 1 GB por archivo")
        + "</div>")

    uploaded = st.file_uploader(
        "Seleccione los archivos PDF",
        type=["pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    if not uploaded:
        st.info("Suba al menos un archivo PDF de 'Autorización de Pago Comisión al Interior del País' para comenzar.")
        _md(FOOTER_HTML)
        return

    # ── Procesamiento ────────────────────────────────────────────────────────
    all_rows = []
    all_warnings = []

    progress = st.progress(0, text="Procesando archivos…")
    for i, f in enumerate(uploaded):
        persons, warns = extract_commission_data(f.read(), filename=f.name)
        all_rows.extend(persons)
        all_warnings.extend(warns)
        progress.progress((i + 1) / len(uploaded), text=f"Procesado: {f.name}")
    progress.empty()

    if all_warnings:
        with st.expander(f"⚠️ Advertencias ({len(all_warnings)})", expanded=True):
            for w in all_warnings:
                st.warning(w)

    if not all_rows:
        st.error("No se pudo extraer información de los archivos cargados.")
        _md(FOOTER_HTML)
        return

    df = pd.DataFrame(all_rows, columns=COLUMNS)
    st.success(f"✅ {len(df)} registro(s) extraído(s) de {len(uploaded)} archivo(s).")

    # ── Resumen ───────────────────────────────────────────────────────────────
    _md(f'<div style="{SECTION_STYLE}">'
        + section_header("📊", "Resumen por solicitud", "Suma de valores individuales = total de la comisión")
        + "</div>")

    summary_rows = []
    for sol_no, group in df.groupby("Solicitud de Comisión No.", sort=False):
        def to_num(v):
            try:
                return float(str(v).replace(".", "").replace(",", "."))
            except Exception:
                return 0.0
        total = sum(to_num(v) for v in group["Valor Total a Pagar"])
        summary_rows.append({
            "Solicitud No.": sol_no,
            "Personas": len(group),
            "Valor Total (COP)": f"{total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        })
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    # ── Detalle ───────────────────────────────────────────────────────────────
    _md(f'<div style="{SECTION_STYLE}">'
        + section_header("📋", "Detalle completo", "Una fila por comisionado — todos los campos extraídos del PDF")
        + "</div>")
    st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Descarga ──────────────────────────────────────────────────────────────
    excel_bytes = build_excel(df)
    st.download_button(
        label="⬇️  Descargar Excel",
        data=excel_bytes,
        file_name="comisiones_SIIF.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    _md(FOOTER_HTML)


if __name__ == "__main__":
    main()
